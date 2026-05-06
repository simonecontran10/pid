"""
apply_sots_overrides.py — Applica mapping manuali da data/sots_overrides.xlsx
ai dati PID. Lavora su:
  - data/sortitoutsi_id_lookup.json (aggiunge entry)
  - data/players_main.json, players_all.json, players_static.json (campi sortitoutsi_*)
  - data/photos/players_sots_lookup/ (scarica face)

Excel atteso: 2 colonne (name, sortitoutsi_url). Header in riga 1.
Le righe con cella vuota in url sono saltate.

Uso:
    python3 apply_sots_overrides.py
    python3 apply_sots_overrides.py --dry-run   # solo report, niente modifiche
"""
from __future__ import annotations

import json
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import _bootstrap  # noqa

import requests
from openpyxl import load_workbook

from scraper.config import (
    DATA_DIR,
    PLAYERS_MAIN_FILE,
    PLAYERS_STATIC_FILE,
    USER_AGENTS,
)
from scraper.sortitoutsi import face_url, profile_url

OVERRIDES_FILE = DATA_DIR / "sots_overrides.xlsx"
LOOKUP_FILE = DATA_DIR / "sortitoutsi_id_lookup.json"
PLAYERS_ALL_FILE = DATA_DIR / "players_all.json"
PHOTOS_DIR = DATA_DIR / "photos" / "players_sots_lookup"
PHOTOS_DIR.mkdir(parents=True, exist_ok=True)

DRY_RUN = "--dry-run" in sys.argv


def extract_sots_id(url: str) -> int | None:
    """Estrae l'ID numerico da un URL sortitoutsi person."""
    m = re.search(r"/person/(\d+)", url or "")
    return int(m.group(1)) if m else None


def download_face(sots_id: int, session: requests.Session) -> Path | None:
    """Scarica la face da sortitoutsi CDN. Ritorna il path locale o None."""
    out = PHOTOS_DIR / f"{sots_id}.png"
    if out.exists() and out.stat().st_size > 0:
        return out
    url = f"https://sortitoutsi.b-cdn.net/uploads/face/face_{sots_id}.png"
    try:
        r = session.get(url, headers={"User-Agent": USER_AGENTS[0]}, timeout=15)
        if r.status_code == 200 and r.content and len(r.content) > 200:
            out.write_bytes(r.content)
            return out
    except Exception as e:
        print(f"    [face dl error {sots_id}] {e}")
    return None


def main():
    if not OVERRIDES_FILE.exists():
        print(f"ERRORE: {OVERRIDES_FILE} non esiste")
        return

    # Carica Excel
    wb = load_workbook(OVERRIDES_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0] or not row[1]:
            continue
        name = str(row[0]).strip()
        url = str(row[1]).strip()
        sots_id = extract_sots_id(url)
        if not sots_id:
            print(f"  [riga {i}] skip — URL non valido: {url!r}")
            continue
        rows.append({"row": i, "name": name, "url": url, "sots_id": sots_id})

    print(f"Lette {len(rows)} righe valide da {OVERRIDES_FILE.name}")

    # Carica players_main
    players = json.loads(PLAYERS_MAIN_FILE.read_text(encoding="utf-8"))
    by_name = {p.get("full_name"): p for p in players}

    # Carica lookup esistente
    lookup = json.loads(LOOKUP_FILE.read_text(encoding="utf-8")) if LOOKUP_FILE.exists() else {}

    # Match
    matched = []
    not_found = []
    for r in rows:
        p = by_name.get(r["name"])
        if not p:
            not_found.append(r)
            continue
        r["tm_player_id"] = p["tm_player_id"]
        r["existing_sots_id"] = p.get("sortitoutsi_person_id")
        matched.append(r)

    print(f"\n=== Match ===")
    print(f"  trovati:  {len(matched)}/{len(rows)}")
    print(f"  no-match: {len(not_found)}/{len(rows)}")
    if not_found:
        print(f"\n  Nomi non trovati nel database:")
        for r in not_found:
            print(f"    riga {r['row']}: {r['name']!r}")

    if not matched:
        return

    # Applica
    print(f"\n=== Apply ===")
    session = requests.Session()
    n_new = 0
    n_overwrite = 0
    n_face_dl = 0
    for r in matched:
        tm_id = r["tm_player_id"]
        sots_id = r["sots_id"]
        existing = r["existing_sots_id"]
        action = "+" if existing is None else ("=" if existing == sots_id else "OVERWRITE")
        if existing is not None and existing != sots_id:
            n_overwrite += 1
        elif existing is None:
            n_new += 1
        print(f"  {action} {r['name']:<35}  tm={tm_id}  sots={sots_id}  (era {existing})")

        if DRY_RUN:
            continue

        # Scarica foto
        face_path = download_face(sots_id, session)
        if face_path:
            n_face_dl += 1

        # Aggiorna lookup
        lookup[str(tm_id)] = {
            "sots_id": sots_id,
            "name": r["name"],
            "match_score": 1.0,
            "matched_via": "manual_override",
            "face_local": f"photos/players_sots_lookup/{sots_id}.png" if face_path else None,
        }
        time.sleep(0.5)

    if DRY_RUN:
        print("\n[DRY RUN] niente scritto su disco")
        return

    # Salva lookup
    LOOKUP_FILE.write_text(json.dumps(lookup, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n  saved lookup ({len(lookup)} entries)")

    # Aggiorna players_main / players_all / players_static
    matched_by_tm = {r["tm_player_id"]: r for r in matched}
    for path in (PLAYERS_MAIN_FILE, PLAYERS_STATIC_FILE, PLAYERS_ALL_FILE):
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        n_changed = 0
        for p in data:
            r = matched_by_tm.get(p.get("tm_player_id"))
            if not r:
                continue
            sots_id = r["sots_id"]
            p["sortitoutsi_person_id"] = sots_id
            p["sortitoutsi_face_url"] = face_url(sots_id)
            p["sortitoutsi_profile_url"] = profile_url(sots_id, p.get("full_name"))
            # Path locale che il frontend usa per la foto FM-style
            p["sortitoutsi_face_local_lookup"] = f"photos/players_sots_lookup/{sots_id}.png"
            n_changed += 1
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  {path.name}: {n_changed} giocatori aggiornati")

    print(f"\n=== Summary ===")
    print(f"  nuovi mapping:  {n_new}")
    print(f"  sovrascritti:   {n_overwrite}")
    print(f"  face scaricate: {n_face_dl}")


if __name__ == "__main__":
    main()
