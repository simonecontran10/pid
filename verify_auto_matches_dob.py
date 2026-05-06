"""
verify_auto_matches_dob.py — Verifica i match automatici confrontando la DOB
Per ogni match candidato (da sots_auto_matched_preview.xlsx):
  - Fetch pagina person SortItOutSi
  - Estrai DOB
  - Confronta con date_of_birth di players_main.json
  - Solo match con DOB esatta vengono confermati
Output:
  - sots_auto_matched_confirmed.xlsx (DOB match -> applicabili)
  - sots_dob_mismatch.xlsx (DOB diversa -> da verificare manualmente)
"""
from __future__ import annotations
import json
import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import _bootstrap  # noqa

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

from scraper.config import USER_AGENTS, SLEEP_BETWEEN_REQUESTS

DATA_DIR = Path("data")
PREVIEW_XLSX = DATA_DIR / "sots_auto_matched_preview.xlsx"
PLAYERS_MAIN = DATA_DIR / "players_main.json"
CONFIRMED_XLSX = DATA_DIR / "sots_auto_matched_confirmed.xlsx"
MISMATCH_XLSX = DATA_DIR / "sots_dob_mismatch.xlsx"


def fetch_dob(sots_id: int, slug: str, session: requests.Session):
    """Fetch pagina person, estrai DOB (YYYY-MM-DD)."""
    url = f"https://sortitoutsi.net/football-manager-2026/person/{sots_id}/{slug}"
    try:
        r = session.get(url, headers={"User-Agent": USER_AGENTS[0]}, timeout=15)
        if r.status_code != 200:
            return None, f"HTTP {r.status_code}"
        # Pattern: <dt>DOB</dt><dd>YYYY-MM-DD</dd>
        m = re.search(r"DOB</dt>\s*<dd[^>]*>(\d{4}-\d{2}-\d{2})</dd>", r.text)
        if m:
            return m.group(1), None
        return None, "DOB pattern not found"
    except Exception as e:
        return None, f"err: {e}"


def main():
    # Carica preview
    wb = load_workbook(PREVIEW_XLSX, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Map tm_id -> dob da players_main
    players = json.loads(PLAYERS_MAIN.read_text(encoding="utf-8"))
    dob_by_tm = {p["tm_player_id"]: str(p.get("date_of_birth", ""))[:10] for p in players}

    print(f"Verifica DOB per {len(rows)} match candidati...")

    session = requests.Session()
    confirmed = []
    mismatched = []
    no_dob = []

    for i, row in enumerate(rows, 1):
        name, tm_id, club, score, why, sots_id, url = row
        slug = url.rsplit("/", 1)[-1] if url else ""
        tm_dob = dob_by_tm.get(tm_id, "")
        sots_dob, err = fetch_dob(sots_id, slug, session)
        time.sleep(SLEEP_BETWEEN_REQUESTS)

        if not sots_dob:
            no_dob.append((name, tm_id, sots_id, url, tm_dob, err))
            print(f"  [{i}/{len(rows)}] {name:<30}  ?? {err}")
            continue

        if sots_dob == tm_dob:
            confirmed.append((name, tm_id, club, score, why, sots_id, url, tm_dob))
            print(f"  [{i}/{len(rows)}] {name:<30}  OK   tm={tm_dob} sots={sots_dob}")
        else:
            mismatched.append((name, tm_id, club, score, why, sots_id, url, tm_dob, sots_dob))
            print(f"  [{i}/{len(rows)}] {name:<30}  MIS  tm={tm_dob} sots={sots_dob}")

    print()
    print(f"=== Risultati ===")
    print(f"  DOB confermata:     {len(confirmed)}")
    print(f"  DOB mismatch:       {len(mismatched)}")
    print(f"  DOB non ottenibile: {len(no_dob)}")

    # Salva confirmed
    wb = Workbook()
    ws = wb.active
    ws.title = "confirmed"
    ws.append(["name", "tm_player_id", "club_name", "score", "why", "sots_id", "url", "dob"])
    for r in confirmed:
        ws.append(list(r))
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["G"].width = 60
    wb.save(CONFIRMED_XLSX)

    # Salva mismatch
    wb = Workbook()
    ws = wb.active
    ws.title = "mismatch"
    ws.append(["name", "tm_player_id", "club_name", "score", "why", "sots_id", "url", "tm_dob", "sots_dob"])
    for r in mismatched + [(n, t, "?", "?", "?", s, u, d, "n/a") for n, t, s, u, d, _ in no_dob]:
        ws.append(list(r))
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["G"].width = 60
    wb.save(MISMATCH_XLSX)

    print(f"\nSalvato {CONFIRMED_XLSX.name} (da applicare con apply_sots_overrides.py)")
    print(f"Salvato {MISMATCH_XLSX.name} (da verificare manualmente)")


if __name__ == "__main__":
    main()
