"""apply_more_matches.py — Applica match da sots_more_matches_confirmed.xlsx."""
from __future__ import annotations
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import _bootstrap  # noqa

import requests
from openpyxl import load_workbook

from scraper.config import (
    DATA_DIR, PLAYERS_MAIN_FILE, PLAYERS_STATIC_FILE, USER_AGENTS,
)
from scraper.sortitoutsi import face_url, profile_url

CONFIRMED_FILE = DATA_DIR / "sots_more_matches_confirmed.xlsx"
LOOKUP_FILE = DATA_DIR / "sortitoutsi_id_lookup.json"
PLAYERS_ALL_FILE = DATA_DIR / "players_all.json"
PHOTOS_DIR = DATA_DIR / "photos" / "players_sots_lookup"
PHOTOS_DIR.mkdir(parents=True, exist_ok=True)


def download_face(sots_id, session):
    out = PHOTOS_DIR / f"{sots_id}.png"
    if out.exists() and out.stat().st_size > 0:
        return out
    url = f"https://sortitoutsi.b-cdn.net/uploads/face/face_{sots_id}.png"
    try:
        r = session.get(url, headers={"User-Agent": USER_AGENTS[0]}, timeout=15)
        if r.status_code == 200 and r.content and len(r.content) > 200:
            out.write_bytes(r.content)
            return out
    except Exception as e:
        print(f"  [face err {sots_id}] {e}")
    return None


def main():
    wb = load_workbook(CONFIRMED_FILE, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        # Schema: name, tm_player_id, club_name, score, why, sots_id, url, dob
        rows.append({
            "name": row[0],
            "tm_id": int(row[1]),
            "sots_id": int(row[5]),
        })

    print(f"Da applicare: {len(rows)} match confermati")

    lookup = json.loads(LOOKUP_FILE.read_text(encoding="utf-8")) if LOOKUP_FILE.exists() else {}

    session = requests.Session()
    n_face_dl = 0
    n_face_cached = 0

    for i, r in enumerate(rows, 1):
        face = download_face(r["sots_id"], session)
        if face:
            if face.stat().st_mtime > time.time() - 30:  # appena creato
                n_face_dl += 1
            else:
                n_face_cached += 1
        lookup[str(r["tm_id"])] = {
            "sots_id": r["sots_id"],
            "name": r["name"],
            "match_score": 1.0,
            "matched_via": "cross_club_dob_confirmed",
            "face_local": f"photos/players_sots_lookup/{r['sots_id']}.png" if face else None,
        }
        if i % 100 == 0 or i == len(rows):
            print(f"  [{i}/{len(rows)}] face_dl={n_face_dl} cached={n_face_cached}")
        time.sleep(0.2)

    LOOKUP_FILE.write_text(json.dumps(lookup, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nlookup salvato ({len(lookup)} entries)")

    # Aggiorna players_main / players_all / players_static
    matched_by_tm = {r["tm_id"]: r for r in rows}
    for path in (PLAYERS_MAIN_FILE, PLAYERS_STATIC_FILE, PLAYERS_ALL_FILE):
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        n_changed = 0
        for p in data:
            r = matched_by_tm.get(p.get("tm_player_id"))
            if not r:
                continue
            sid = r["sots_id"]
            p["sortitoutsi_person_id"] = sid
            p["sortitoutsi_face_url"] = face_url(sid)
            p["sortitoutsi_profile_url"] = profile_url(sid, p.get("full_name"))
            p["sortitoutsi_face_local_lookup"] = f"photos/players_sots_lookup/{sid}.png"
            n_changed += 1
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  {path.name}: {n_changed} giocatori aggiornati")

    print(f"\nFace: {n_face_dl} scaricate, {n_face_cached} già in cache")


if __name__ == "__main__":
    main()
