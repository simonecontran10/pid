"""Filtra i match veri da sots_more_matches_dob_mismatch.xlsx.

PROBLEMA: il TM scraper estrae DOB tronche per giocatori Primavera (es. "2007 (~ 19")
quando la pagina TM nasconde il giorno/mese (privacy minorenni). Lo script
find_more_sots_matches.py marca questi come DOB mismatch.

SOLUZIONE: per ogni candidato, accetta come VERO MATCH se:
  - Anno TM (estratto come int dai primi 4 char di tm_dob) == anno SOTS (primi 4 char)
  - Slug TM == slug SOTS (perfetto, già condizione di find_more_*)
  - Club TM e SOTS sono compatibili (mappa Primavera→club madre)

Risultato: file sots_more_matches_confirmed.xlsx pronto per apply_more_matches.py
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from collections import defaultdict

DATA_DIR = Path("data")
MISMATCH_XLSX = DATA_DIR / "sots_more_matches_dob_mismatch.xlsx"
CONFIRMED_XLSX = DATA_DIR / "sots_more_matches_confirmed.xlsx"

# Mappa club madre → varianti accettabili
# Primavera → club Serie A/B
def club_root(name: str) -> str:
    """Estrae 'parola madre' del club (rimuove Primavera, U23, ecc.)."""
    if not name:
        return ""
    n = name.lower()
    for suffix in [" primavera", " u23", " u19", " u18", " ii", " futuro"]:
        n = n.replace(suffix, "")
    # Rimuovi prefissi/suffissi standard club
    for prefix in ["ac ", "as ", "fc ", "us ", "ss ", "ssc ", "uc ", "acf ", "sef ", "ucf "]:
        if n.startswith(prefix):
            n = n[len(prefix):]
    for suffix in [" calcio", " 1909", " 1907", " 1912", " 1919", " 1920"]:
        n = n.replace(suffix, "")
    # Rimuove date alla fine "1898", "1903", ecc.
    import re
    n = re.sub(r'\s+\d{4}$', '', n)
    return n.strip()


def main():
    if not MISMATCH_XLSX.exists():
        print(f"❌ {MISMATCH_XLSX} non esiste"); return

    wb = load_workbook(MISMATCH_XLSX, data_only=True)
    ws = wb.active

    # Header: name, tm_player_id, tm_club, tm_dob, sots_dob, sots_id, slug, sots_club, url
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append({
            "name": row[0],
            "tm_id": int(row[1]),
            "tm_club": row[2] or "",
            "tm_dob": str(row[3] or ""),
            "sots_dob": str(row[4] or ""),
            "sots_id": int(row[5]),
            "slug": row[6] or "",
            "sots_club": row[7] or "",
            "url": row[8] or "",
        })

    print(f"Righe da analizzare: {len(rows)}")

    # Step 1: estrai anno TM e SOTS
    def extract_year(s: str) -> int:
        if not s or len(s) < 4:
            return 0
        try:
            return int(s[:4])
        except ValueError:
            return 0

    # Step 2: raggruppa per tm_id (per gestire omonimi)
    by_tm = defaultdict(list)
    for r in rows:
        by_tm[r["tm_id"]].append(r)

    confirmed = []
    skipped_year = []
    skipped_club = []
    skipped_ambiguous = []

    for tm_id, candidates in by_tm.items():
        # Filtra solo candidati con anno coincidente
        valid_year = [c for c in candidates if extract_year(c["tm_dob"]) == extract_year(c["sots_dob"])]
        if not valid_year:
            skipped_year.extend(candidates)
            continue

        # Tra i candidati con anno valido, prendi solo quelli col club compatibile
        valid_club = []
        for c in valid_year:
            tm_root = club_root(c["tm_club"])
            sots_root = club_root(c["sots_club"])
            # Match esatto OR substring (per gestire variazioni di nome)
            if tm_root == sots_root or tm_root in sots_root or sots_root in tm_root:
                valid_club.append(c)
        
        if not valid_club:
            # Anno coincide ma club no — sospetto omonimia, scarta
            skipped_club.extend(valid_year)
            continue

        if len(valid_club) > 1:
            # Più candidati anche dopo filtro club: ambiguo, scarta tutti
            skipped_ambiguous.extend(valid_club)
            continue

        # 1 candidato valido = vero match
        confirmed.append(valid_club[0])

    print()
    print(f"✅ Confermati (anno+club): {len(confirmed)}")
    print(f"❌ Scartati (anno diverso): {len(skipped_year)}")
    print(f"❌ Scartati (club incompatibile): {len(skipped_club)}")
    print(f"❌ Scartati (ambigui, omonimi): {len(skipped_ambiguous)}")

    if skipped_year:
        print(f"\n--- Esempi scartati per anno diverso ---")
        for c in skipped_year[:5]:
            print(f"  {c['name']:30} TM:{c['tm_dob']:15} SOTS:{c['sots_dob']:12} ({c['sots_club']})")

    if skipped_ambiguous:
        print(f"\n--- Esempi ambigui (omonimi) ---")
        seen = set()
        for c in skipped_ambiguous:
            if c['tm_id'] not in seen:
                same = [x for x in skipped_ambiguous if x['tm_id'] == c['tm_id']]
                print(f"  {c['name']:30} TM_club:{c['tm_club']}")
                for s in same:
                    print(f"      → SOTS: {s['sots_club']:30} dob:{s['sots_dob']}")
                seen.add(c['tm_id'])

    # Scrivi file confirmed
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "confirmed"
    ws_out.append(['name', 'tm_player_id', 'club_name', 'score', 'why', 'sots_id', 'url', 'dob'])
    for c in confirmed:
        ws_out.append([
            c['name'],
            c['tm_id'],
            c['tm_club'],
            1.0,
            'year_club_match',
            c['sots_id'],
            c['url'],
            c['sots_dob'],
        ])
    wb_out.save(CONFIRMED_XLSX)
    print(f"\n✅ Salvato {CONFIRMED_XLSX} ({len(confirmed)+1} righe header+dati)")
    print(f"\nProssimo step: lancia python3 apply_more_matches.py")


if __name__ == "__main__":
    main()
